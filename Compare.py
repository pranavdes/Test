import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import os
import hashlib
from difflib import SequenceMatcher
import argparse
import sys
import logging
from typing import Dict, List, Tuple
from tqdm import tqdm
from concurrent.futures import ProcessPoolExecutor
import multiprocessing
import json
import csv

# Set up logging configuration
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def unmerge_cells(worksheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
    """
    Unmerge all merged cells in the worksheet and copy the merged value to each cell.
    
    Args:
        worksheet (openpyxl.worksheet.worksheet.Worksheet): The worksheet to process.
    """
    try:
        # Get all merged cell ranges
        merged_ranges = list(worksheet.merged_cells.ranges)
        for merged_range in merged_ranges:
            # Unmerge the cells
            worksheet.unmerge_cells(str(merged_range))
            # Get the value from the top-left cell of the merged range
            value = worksheet.cell(merged_range.min_row, merged_range.min_col).value
            # Copy the value to all cells in the previously merged range
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    worksheet.cell(row, col, value)
    except Exception as e:
        logger.error(f"Error in unmerging cells: {e}")
        raise

def get_function_details(workbook: openpyxl.workbook.workbook.Workbook, sheet_name: str, key_column: str) -> Dict[str, Dict[str, str]]:
    """
    Extract function details from the specified sheet.
    
    Args:
        workbook (openpyxl.workbook.workbook.Workbook): The workbook containing the function details.
        sheet_name (str): The name of the sheet containing function details.
        key_column (str): The name of the column containing the function IDs.
    
    Returns:
        Dict[str, Dict[str, str]]: A dictionary mapping function IDs to their details.
    """
    function_details = {}
    try:
        if sheet_name not in workbook.sheetnames:
            logger.warning(f"Sheet '{sheet_name}' not found in workbook. Function details will be empty.")
            return function_details
        
        sheet = workbook[sheet_name]
        headers = [cell.value for cell in sheet[1]]
        key_index = headers.index(key_column)
        name_index = headers.index('Function Name')
        owner_index = headers.index('Owner')
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[key_index]:
                function_details[str(row[key_index])] = {'name': row[name_index], 'owner': row[owner_index]}
    except Exception as e:
        logger.error(f"Error in getting function details: {e}")
        raise
    return function_details

def generate_hash(values: List[str]) -> str:
    """
    Generate a hash value based on a list of values.

    Args:
        values (List[str]): List of values to hash.

    Returns:
        str: A hash string representing the combined values.
    """
    hash_input = '|'.join([str(v) for v in values])
    return hashlib.blake2b(hash_input.encode('utf-8'), digest_size=16).hexdigest()

def compare_strings(s1: str, s2: str) -> float:
    """
    Compare two strings and return a similarity ratio using SequenceMatcher.

    Args:
        s1 (str): The first string to compare.
        s2 (str): The second string to compare.

    Returns:
        float: A similarity ratio between 0 and 1.
    """
    s1 = str(s1).strip().lower()
    s2 = str(s2).strip().lower()
    return SequenceMatcher(None, s1, s2).ratio()

def categorize_change(ratio: float, minor_threshold: float, major_threshold: float) -> str:
    """
    Categorize the type of change based on the similarity ratio.

    Args:
        ratio (float): The similarity ratio between 0 and 1.
        minor_threshold (float): Threshold above which a change is considered minor.
        major_threshold (float): Threshold above which a change is considered major.

    Returns:
        str: The category of change ('No change', 'Minor change', 'Major change', 'Substantial change').
    """
    if ratio == 1:
        return 'No change'
    elif ratio >= minor_threshold:
        return 'Minor change'
    elif ratio >= major_threshold:
        return 'Major change'
    else:
        return 'Substantial change'

def compare_chunks(chunk1: pd.DataFrame, chunk2: pd.DataFrame, sheet_name: str, minor_threshold: float, major_threshold: float, function_details: Dict[str, Dict[str, str]]) -> List[Tuple]:
    """
    Compare two chunks of data and return the comparison results.
    
    Args:
        chunk1 (pd.DataFrame): DataFrame chunk from the first sheet.
        chunk2 (pd.DataFrame): DataFrame chunk from the second sheet.
        sheet_name (str): Name of the sheet being compared.
        minor_threshold (float): Threshold for minor changes.
        major_threshold (float): Threshold for major changes.
        function_details (Dict[str, Dict[str, str]]): Dictionary containing function details.
    
    Returns:
        List[Tuple]: List of comparison results.
    """
    results = []
    
    # Ensure index is treated as strings
    chunk1.index = chunk1.index.astype(str)
    chunk2.index = chunk2.index.astype(str)
    
    # Combine DataFrames for comparison
    combined = pd.concat([chunk1, chunk2], keys=['source1', 'source2'])
    
    for col in chunk1.columns:
        grouped = combined.groupby(level=1)[col]
        for idx, group in grouped:
            if len(group) == 2:  # Value exists in both sources
                val1, val2 = group.values
                if pd.isna(val1) and pd.isna(val2):
                    continue
                ratio = compare_strings(str(val1), str(val2))
                change_type = categorize_change(ratio, minor_threshold, major_threshold)
                
                # Get function details
                function_id = str(chunk1.index[chunk1.index == idx][0])
                function_name = function_details.get(function_id, {}).get('name', '')
                owner = function_details.get(function_id, {}).get('owner', '')
                
                cell1 = f"{get_column_letter(chunk1.columns.get_loc(col) + 1)}{idx} ({col})"
                cell2 = f"{get_column_letter(chunk2.columns.get_loc(col) + 1)}{idx} ({col})"
                
                results.append((function_id, function_name, owner, sheet_name, cell1, val1, cell2, val2, change_type))
            elif len(group) == 1:  # Value exists in only one source
                source = group.index[0][0]
                val = group.values[0]
                if pd.isna(val):
                    continue
                
                # Get function details
                function_id = str(idx)
                function_name = function_details.get(function_id, {}).get('name', '')
                owner = function_details.get(function_id, {}).get('owner', '')
                
                if source == 'source1':
                    cell = f"{get_column_letter(chunk1.columns.get_loc(col) + 1)}{idx} ({col})"
                    results.append((function_id, function_name, owner, sheet_name, cell, val, '', '', 'Cell deleted'))
                else:
                    cell = f"{get_column_letter(chunk2.columns.get_loc(col) + 1)}{idx} ({col})"
                    results.append((function_id, function_name, owner, sheet_name, '', '', cell, val, 'Cell added'))
    
    return results

def compare_sheets(sheet1: pd.DataFrame, sheet2: pd.DataFrame, sheet_name: str, minor_threshold: float, major_threshold: float, chunk_size: int, function_details: Dict[str, Dict[str, str]]) -> List[Tuple]:
    """
    Compare two sheets and return the comparison results.

    Args:
        sheet1 (pd.DataFrame): DataFrame of the first sheet.
        sheet2 (pd.DataFrame): DataFrame of the second sheet.
        sheet_name (str): Name of the sheet being compared.
        minor_threshold (float): Threshold for minor changes.
        major_threshold (float): Threshold for major changes.
        chunk_size (int): Number of rows to process at a time.
        function_details (Dict[str, Dict[str, str]]): Dictionary containing function details.

    Returns:
        List[Tuple]: List of comparison results.
    """
    results = []
    
    # Process the sheets in chunks
    for start in range(0, len(sheet1), chunk_size):
        end = start + chunk_size
        chunk1 = sheet1.iloc[start:end]
        chunk2 = sheet2.iloc[start:end]
        
        chunk_results = compare_chunks(chunk1, chunk2, sheet_name, minor_threshold, major_threshold, function_details)
        results.extend(chunk_results)
    
    return results

def process_sheet(args):
    """
    Process a single sheet for comparison.

    Args:
        args: Tuple containing (sheet_name, file1_path, file2_path, minor_threshold, major_threshold, chunk_size, function_details)

    Returns:
        Tuple: (sheet_name, comparison_results)
    """
    sheet_name, file1_path, file2_path, minor_threshold, major_threshold, chunk_size, function_details = args
    try:
        sheet1 = pd.read_excel(file1_path, sheet_name=sheet_name)
        sheet2 = pd.read_excel(file2_path, sheet_name=sheet_name)
        results = compare_sheets(sheet1, sheet2, sheet_name, minor_threshold, major_threshold, chunk_size, function_details)
        return sheet_name, results
    except Exception as e:
        logger.error(f"Error processing sheet {sheet_name}: {e}")
        return sheet_name, []

def compare_excel_files(
    file1_path: str,
    file2_path: str,
    output_path: str,
    minor_threshold: float,
    major_threshold: float,
    ignore_sheets: List[str],
    chunk_size: int,
    num_processes: int,
    output_format: str
) -> None:
    """
    Main function to compare two Excel files and generate a comparison report.
    
    Args:
        file1_path (str): Path to the first Excel file.
        file2_path (str): Path to the second Excel file.
        output_path (str): Path to save the output file.
        minor_threshold (float): Threshold for minor changes.
        major_threshold (float): Threshold for major changes.
        ignore_sheets (List[str]): List of sheet names to ignore.
        chunk_size (int): Number of rows to process at a time.
        num_processes (int): Number of processes to use for parallel processing.
        output_format (str): Format of the output file ('excel', 'csv', or 'json').
    """
    try:
        # Load workbooks
        wb1 = openpyxl.load_workbook(file1_path, read_only=True)
        wb2 = openpyxl.load_workbook(file2_path, read_only=True)
        
        sheets_to_compare = [sheet for sheet in wb1.sheetnames if sheet in wb2.sheetnames and sheet not in ignore_sheets]
        
        # Get function details
        function_details_sheet = 'Core OCIR Data'  # Update if different
        function_details = get_function_details(wb1, function_details_sheet, 'Function ID')
        
        # Close workbooks to free up memory
        wb1.close()
        wb2.close()
        
    except Exception as e:
        logger.error(f"Error loading workbooks: {e}")
        sys.exit(1)

    # Prepare arguments for multiprocessing
    args_list = [(sheet, file1_path, file2_path, minor_threshold, major_threshold, chunk_size, function_details) for sheet in sheets_to_compare]

    # Process sheets in parallel
    with ProcessPoolExecutor(max_workers=num_processes) as executor:
        results = list(tqdm(executor.map(process_sheet, args_list), total=len(args_list), desc="Processing sheets"))

    # Aggregate results
    all_results = [result for sheet_results in results for result in sheet_results]

    # Generate output based on the specified format
    if output_format == 'excel':
        generate_excel_output(all_results, output_path, file1_path, file2_path)
    elif output_format == 'csv':
        generate_csv_output(all_results, output_path)
    elif output_format == 'json':
        generate_json_output(all_results, output_path)
    else:
        logger.error(f"Unsupported output format: {output_format}")
        sys.exit(1)

    logger.info(f"Comparison report saved to {output_path}")

def generate_excel_output(results: List[Tuple], output_path: str, file1_path: str, file2_path: str) -> None:
    """
    Generate an Excel output file with the comparison results.
    
    Args:
        results (List[Tuple]): List of comparison results.
        output_path (str): Path to save the output Excel file.
        file1_path (str): Path of the first input Excel file.
        file2_path (str): Path of the second input Excel file.
    """
    wb_output = openpyxl.Workbook()
    ws_output = wb_output.active
    ws_output.title = 'Comparison'

    # Define colors for different types of changes
    colors = {
        'Cell added': 'C6EFCE',
        'Cell deleted': 'FFC7CE',
        'Minor change': 'FFEB9C',
        'Major change': 'FFD966',
        'Substantial change': 'F4B084',
        'Moved with no change': 'D9E1F2',
        'No change': 'FFFFFF'
    }

    # Set up the header row
    headers = ['Sr. No', 'Function ID', 'Function Name', 'Owner', 'Sheet Name', 'Source 1 Cell', 'Source 1 Value', 'Source 2 Cell', 'Source 2 Value', 'Change Summary']
    ws_output.append(headers)

    # Format the header row
    for col in range(1, len(headers) + 1):
        cell = ws_output.cell(row=1, column=col)
        cell.font = Font(bold=True, color='000000')
        cell.fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Add source file details
    ws_output['L2'] = 'Source 1:'
    ws_output['M2'] = os.path.basename(file1_path)
    ws_output['L3'] = 'Source 2:'
    ws_output['M3'] = os.path.basename(file2_path)

    # Add comparison results
    for idx, result in enumerate(results, start=2):
        function_id, function_name, owner, sheet_name, cell1, val1, cell2, val2, change_type = result
        ws_output.append([
            idx - 1,  # Sr. No
            function_id,
            function_name,
            owner,
            sheet_name,
            cell1,
            val1,
            cell2,
            val2,
            change_type
        ])
        
        # Apply color formatting
        for col in range(1, 11):
            ws_output.cell(row=idx, column=col).fill = PatternFill(
                start_color=colors.get(change_type, 'FFFFFF'),
                end_color=colors.get(change_type, 'FFFFFF'),
                fill_type='solid'
            )

    # Add color legends
    ws_output['L5'] = 'Color Legend'
    ws_output['L5'].font = Font(bold=True)
    for i, (change_type, color) in enumerate(colors.items(), start=6):
        cell = ws_output.cell(row=i, column=12, value=change_type)
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Adjust column widths
    for col in ws_output.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min((max_length + 2), 50)  # Cap width at 50
        ws_output.column_dimensions[column].width = adjusted_width

    # Save the output workbook
    wb_output.save(output_path)

def generate_csv_output(results: List[Tuple], output_path: str) -> None:
    """
    Generate a CSV output file with the comparison results.
    
    Args:
        results (List[Tuple]): List of comparison results.
        output_path (str): Path to save the output CSV file.
    """
    with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(['Sr. No', 'Function ID', 'Function Name', 'Owner', 'Sheet Name', 'Source 1 Cell', 'Source 1 Value', 'Source 2 Cell', 'Source 2 Value', 'Change Summary'])
        
        for idx, result in enumerate(results, start=1):
            function_id, function_name, owner, sheet_name, cell1, val1, cell2, val2, change_type = result
            writer.writerow([idx, function_id, function_name, owner, sheet_name, cell1, val1, cell2, val2, change_type])

def generate_json_output(results: List[Tuple], output_path: str) -> None:
    """
    Generate a JSON output file with the comparison results.
    
    Args:
        results (List[Tuple]): List of comparison results.
        output_path (str): Path to save the output JSON file.
    """
    json_results = []
    for idx, result in enumerate(results, start=1):
        function_id, function_name, owner, sheet_name, cell1, val1, cell2, val2, change_type = result
        json_results.append({
            'Sr. No': idx,
            'Function ID': function_id,
            'Function Name': function_name,
            'Owner': owner,
            'Sheet Name': sheet_name,
            'Source 1 Cell': cell1,
            'Source 1 Value': val1,
            'Source 2 Cell': cell2,
            'Source 2 Value': val2,
            'Change Summary': change_type
        })
    
    with open(output_path, 'w', encoding='utf-8') as jsonfile:
        json.dump(json_results, jsonfile, indent=2)

def main():
    """
    Main function to handle command-line arguments and initiate the comparison process.
    """
    parser = argparse.ArgumentParser(description='Compare two Excel files and generate a comparison report.')
    parser.add_argument('-f1', '--file1', type=str, help='Path to the first Excel file.')
    parser.add_argument('-f2', '--file2', type=str, help='Path to the second Excel file.')
    parser.add_argument('-o', '--output', type=str, help='Path to the output file.')
    parser.add_argument('-mth', '--minor_threshold', type=float, default=0.8, help='Threshold for minor changes (default: 0.8).')
    parser.add_argument('-majth', '--major_threshold', type=float, default=0.5, help='Threshold for major changes (default: 0.5).')
    parser.add_argument('-is', '--ignore_sheets', nargs='*', default=[], help='Sheets to ignore during comparison.')
    parser.add_argument('-cs', '--chunk_size', type=int, default=1000, help='Number of rows to process at a time (default: 1000).')
    parser.add_argument('-p', '--processes', type=int, default=multiprocessing.cpu_count(), help='Number of processes to use (default: number of CPU cores).')
    parser.add_argument('-f', '--format', choices=['excel', 'csv', 'json'], default='excel', help='Output format (default: excel).')
    parser.add_argument('-v', '--verbose', action='store_true', help='Enable verbose logging.')

    args = parser.parse_args()

    # Set logging level
    if args.verbose:
        logger.setLevel(logging.DEBUG)

    # Validate thresholds
    if not 0 <= args.minor_threshold <= 1 or not 0 <= args.major_threshold <= 1:
        logger.error("Thresholds must be between 0 and 1.")
        sys.exit(1)
    if args.minor_threshold <= args.major_threshold:
        logger.error("Minor threshold must be greater than major threshold.")
        sys.exit(1)

    # Set default file paths if not provided
    file1_path = args.file1 or 'source1.xlsx'
    file2_path = args.file2 or 'source2.xlsx'
    output_path = args.output or 'comparison_output.xlsx'

    # Validate file paths
    if not os.path.exists(file1_path):
        logger.error(f"File not found: {file1_path}")
        sys.exit(1)
    if not os.path.exists(file2_path):
        logger.error(f"File not found: {file2_path}")
        sys.exit(1)

    # Log configurations
    logger.info(f"Comparing {file1_path} and {file2_path}")
    logger.info(f"Output will be saved to {output_path}")
    logger.info(f"Minor change threshold: {args.minor_threshold}")
    logger.info(f"Major change threshold: {args.major_threshold}")
    logger.info(f"Chunk size: {args.chunk_size}")
    logger.info(f"Number of processes: {args.processes}")
    logger.info(f"Output format: {args.format}")
    if args.ignore_sheets:
        logger.info(f"Ignoring sheets: {', '.join(args.ignore_sheets)}")

    # Run the comparison
    try:
        compare_excel_files(
            file1_path=file1_path,
            file2_path=file2_path,
            output_path=output_path,
            minor_threshold=args.minor_threshold,
            major_threshold=args.major_threshold,
            ignore_sheets=args.ignore_sheets,
            chunk_size=args.chunk_size,
            num_processes=args.processes,
            output_format=args.format
        )
        logger.info("Comparison completed successfully.")
    except Exception as e:
        logger.error(f"An error occurred during comparison: {e}")
        sys.exit(1)

if __name__ == '__main__':
    main()cell.font = Font(bold=True, color='000000')
        cell.fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Add source file details
    ws_output['L2'] = 'Source 1:'
    ws_output['M2'] = os.path.basename(file1_path)
    ws_output['L3'] = 'Source 2:'
    ws_output['M3'] = os.path.basename(file2_path)

    # Add comparison results
    for idx, result in enumerate(results, start=2):
        function_id, function_name, owner, sheet_name, cell1, val1, cell2, val2, change_type = result
        ws_output.append([
            idx - 1,  # Sr. No
            function_id,
            function_name,
            owner,
            sheet_name,
            cell1,
            val1,
            cell2,
            val2,
            change_type
        ])
        
        # Apply color formatting
        for col in range(1, 11):
            ws_output.cell(row=idx, column=col).fill = PatternFill(
                start_color=colors.get(change_type, 'FFFFFF'),
                end_color=colors.get(change_type, 'FFFFFF'),
                fill_type='solid'
            )

    # Add color legends
    ws_output['L5'] = 'Color Legend'
    ws_output['L5'].font = Font(bold=True)
    for i, (change_type, color) in enumerate(colors.items(), start=6):
        cell = ws_output.cell(row=i, column=12, value=change_type)
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Adjust column widths
    for col in ws_output.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min((max_length + 2), 50)  # Cap width at 50
        ws_output.column_dimensions[column].width = adjusted_width

    # Save the output workbook
    wb_output.save(output_path)

def generate_csv_output(results: List[Tuple], output_path: str) -> None:
    """
    Generate a CSV output file with the comparison results.
    
    Args:
        results (List[Tuple]): List of comparison results.
        output_path (str): Path to save the output CSV file.
    """
    with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(['Sr. No', 'Function ID', 'Function Name', 'Owner', 'Sheet Name', 'Source 1 Cell', 'Source 1 Value', 'Source 2 Cell', 'Source 2 Value', 'Change Summary'])
        
        for idx, result in enumerate(results, start=1):
            function_id, function_name, owner, sheet_name, cell1, val1, cell2, val2, change_type = result
            writer.writerow([idx, function_id, function_name, owner, sheet_name, cell1, val1, cell2, val2, change_type])

def generate_json_output(results: List[Tuple], output_path: str) -> None:
    """
    Generate a JSON output file with the comparison results.
    
    Args:
        results (List[Tuple]): List of comparison results.
        output_path (str): Path to save the output JSON file.
    """
    json_results = []
    for idx, result in enumerate(results, start=1):
        function_id, function_name, owner, sheet_name, cell1, val1, cell2, val2, change_type = result
        json_results.append({
            'Sr. No': idx,
            'Function ID': function_id,
            'Function Name': function_name,
            'Owner': owner,
            'Sheet Name': sheet_name,
            'Source 1 Cell': cell1,
            'Source 1 Value': val1,
            'Source 2 Cell': cell2,
            'Source 2 Value': val2,
            'Change Summary': change_type
        })
    
    with open(output_path, 'w', encoding='utf-8') as jsonfile:
        json.dump(json_results, jsonfile, indent=2)

def main():
    """
    Main function to handle command-line arguments and initiate the comparison process.
    """
    parser = argparse.ArgumentParser(description='Compare two Excel files and generate a comparison report.')
    parser.add_argument('-f1', '--file1', type=str, help='Path to the first Excel file.')
    parser.add_argument('-f2', '--file2', type=str, help='Path to the second Excel file.')
    parser.add_argument('-o', '--output', type=str, help='Path to the output file.')
    parser.add_argument('-mth', '--minor_threshold', type=float, default=0.8, help='Threshold for minor changes (default: 0.8).')
    parser.add_argument('-majth', '--major_threshold', type=float, default=0.5, help='Threshold for major changes (default: 0.5).')
    parser.add_argument('-is', '--ignore_sheets', nargs='*', default=[], help='Sheets to ignore during comparison.')
    parser.add_argument('-cs', '--chunk_size', type=int, default=1000, help='Number of rows to process at a time (default: 1000).')
    parser.add_argument('-p', '--processes', type=int, default=multiprocessing.cpu_count(), help='Number of processes to use (default: number of CPU cores).')
    parser.add_argument('-f', '--format', choices=['excel', 'csv', 'json'], default='excel', help='Output format (default: excel).')
    parser.add_argument('-v', '--verbose', action='store_true', help='Enable verbose logging.')

    args = parser.parse_args()

    # Set logging level
    if args.verbose:
        logger.setLevel(logging.DEBUG)

    # Validate thresholds
    if not 0 <= args.minor_threshold <= 1 or not 0 <= args.major_threshold <= 1:
        logger.error("Thresholds must be between 0 and 1.")
        sys.exit(1)
    if args.minor_threshold <= args.major_threshold:
        logger.error("Minor threshold must be greater than major threshold.")
        sys.exit(1)

    # Set default file paths if not provided
    file1_path = args.file1 or 'source1.xlsx'
    file2_path = args.file2 or 'source2.xlsx'
    output_path = args.output or 'comparison_output.xlsx'

    # Validate file paths
    if not os.path.exists(file1_path):
        logger.error(f"File not found: {file1_path}")
        sys.exit(1)
    if not os.path.exists(file2_path):
        logger.error(f"File not found: {file2_path}")
        sys.exit(1)

    # Log configurations
    logger.info(f"Comparing {file1_path} and {file2_path}")
    logger.info(f"Output will be saved to {output_path}")
    logger.info(f"Minor change threshold: {args.minor_threshold}")
    logger.info(f"Major change threshold: {args.major_threshold}")
    logger.info(f"Chunk size: {args.chunk_size}")
    logger.info(f"Number of processes: {args.processes}")
    logger.info(f"Output format: {args.format}")
    if args.ignore_sheets:
        logger.info(f"Ignoring sheets: {', '.join(args.ignore_sheets)}")

    # Run the comparison
    try:
        compare_excel_files(
            file1_path=file1_path,
            file2_path=file2_path,
            output_path=output_path,
            minor_threshold=args.minor_threshold,
            major_threshold=args.major_threshold,
            ignore_sheets=args.ignore_sheets,
            chunk_size=args.chunk_size,
            num_processes=args.processes,
            output_format=args.format
        )
        logger.info("Comparison completed successfully.")
    except Exception as e:
        logger.error(f"An error occurred during comparison: {e}")
        sys.exit(1)

if __name__ == '__main__':
    main()
