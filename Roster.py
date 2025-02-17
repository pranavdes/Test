# Import necessary libraries
import calendar                              # Used for calendar-related operations, like finding number of days in a month.
from datetime import datetime, timedelta     # Used for working with dates and times.
import pandas as pd                          # Pandas is used to manage and process tabular data (like Excel tables).
from pulp import LpProblem, LpMaximize, LpVariable, lpSum, LpBinary, LpInteger, LpStatus  
                                             # PuLP is used for formulating and solving linear programming problems (our ILP).
from openpyxl import load_workbook, Workbook  # Openpyxl is used to read from and write to Excel workbooks.
from openpyxl.styles import PatternFill, Font, Alignment  
                                             # Used to format Excel cells (e.g., making headers bold or coloring them).
from openpyxl.utils import get_column_letter    # Converts numerical column indexes to Excel column letters (e.g., 1 -> 'A').

# ------------------------------
# Helper functions for Excel I/O
# ------------------------------

def get_named_cell_value(wb, cell_name):
    """
    Retrieves the value from a named cell in the Excel workbook.
    A named cell is a single cell that has been given a name in Excel.
    """
    try:
        # Access the defined names in the workbook and iterate through each destination.
        defined_range = wb.defined_names[cell_name]
        for title, coord in defined_range.destinations:
            ws = wb[title]             # Get the worksheet where the named cell is.
            return ws[coord].value     # Return the value of the cell at that coordinate.
    except Exception as e:
        # If any error occurs, raise a message with the cell name.
        raise ValueError(f"Error reading named cell '{cell_name}': {e}")

def get_table_as_df(ws, table_name):
    """
    Reads an Excel Table (ListObject) from the given worksheet into a Pandas DataFrame.
    It assumes that the first row of the table is a header row.
    """
    for tbl in ws._tables:             # Iterate over all tables in the worksheet.
        if tbl.name == table_name:     # Check if the table name matches.
            ref = tbl.ref              # The reference (cell range) of the table.
            cells = ws[ref]            # Get all cells in the table range.
            # Build a list of lists containing cell values for each row.
            data = [[cell.value for cell in row] for row in cells]
            # The first row is assumed to be the header; create a DataFrame using it.
            return pd.DataFrame(data[1:], columns=data[0])
    # If the table is not found, raise an error.
    raise ValueError(f"Table '{table_name}' not found on worksheet '{ws.title}'.")

def get_working_dates(year, month, public_holidays):
    """
    Returns a sorted list of working dates (as datetime objects) for the given month and year.
    Working dates exclude weekends (Saturday and Sunday) and any dates that are in the public_holidays list.
    """
    num_days = calendar.monthrange(year, month)[1]  # Get number of days in the month.
    all_dates = [datetime(year, month, d) for d in range(1, num_days+1)]  # List all dates in the month.
    # Convert public holiday dates to a set of date objects for fast lookup.
    holiday_dates = set(pd.to_datetime(h).date() for h in public_holidays)
    # Filter out weekends (weekday() >= 5 means Saturday=5, Sunday=6) and public holidays.
    working = [dt for dt in all_dates if dt.weekday() < 5 and dt.date() not in holiday_dates]
    return sorted(working)

def parse_days_string(days_str):
    """
    Given a string of days (e.g., "Mon, Wed, Fri" or "Monday, Tuesday"), 
    returns a set of lowercase day abbreviations and full names.
    For example, "Mon, Wed" becomes {"mon", "monday", "wed", "wednesday"}.
    This is used to check if a seat is available on a certain day.
    """
    if not isinstance(days_str, str):
        return set()                   # If input is not a string, return an empty set.
    parts = [p.strip() for p in days_str.split(',')]  # Split the string by commas.
    day_set = set()
    for p in parts:
        p_l = p.lower()              # Convert to lowercase.
        if p_l.startswith("mon"):
            day_set.update(["mon", "monday"])
        elif p_l.startswith("tue"):
            day_set.update(["tue", "tuesday"])
        elif p_l.startswith("wed"):
            day_set.update(["wed", "wednesday"])
        elif p_l.startswith("thu"):
            day_set.update(["thu", "thursday"])
        elif p_l.startswith("fri"):
            day_set.update(["fri", "friday"])
    return day_set

# ------------------------------
# Functions for day descriptor matching
# ------------------------------

def parse_day_descriptor(descriptor):
    """
    Parses a day descriptor string (e.g., "1st Working Tuesday" or "Last Fri")
    and returns a tuple (occurrence, weekday).
    - occurrence: A string like "1st" or "last" indicating which occurrence of the day.
    - weekday: A normalized day string (e.g., "mon" for Monday).
    If the descriptor cannot be parsed, returns (None, None).
    """
    descriptor = descriptor.strip().lower()  # Clean up the descriptor.
    tokens = descriptor.split()              # Split into individual words.
    valid_occurrences = {"1st", "2nd", "3rd", "4th", "5th", "last"}
    valid_short_days = {"mon", "tue", "wed", "thu", "fri"}
    occ, wday = None, None
    for token in tokens:
        if token in valid_occurrences:
            occ = token                     # Set occurrence if token is valid.
        else:
            for day in valid_short_days:
                if token.startswith(day):
                    wday = day            # Set weekday if token starts with a valid day.
                    break
    return (occ, wday) if occ and wday else (None, None)

def is_day_descriptor_match(date_obj, descriptor, working_dates):
    """
    Checks if a given date (date_obj) matches the descriptor (like "1st Tuesday" or "Last Fri")
    based on the list of working dates.
    Returns True if it matches, otherwise False.
    """
    occ, wday = parse_day_descriptor(descriptor)  # Parse the descriptor.
    if not occ or not wday:
        return False
    # Create a mapping from day abbreviations to weekday numbers (Monday=0, ..., Friday=4).
    day_map = {"mon": 0, "tue": 1, "wed": 2, "thu": 3, "fri": 4}
    needed_wd = day_map.get(wday)
    # Check if the date's weekday matches the needed weekday.
    if date_obj.weekday() != needed_wd:
        return False
    # Collect all working dates in the same month and year that have the same weekday.
    same_wd = [d for d in working_dates if d.year == date_obj.year and d.month == date_obj.month and d.weekday() == needed_wd]
    same_wd.sort()  # Sort the list in chronological order.
    if occ == "last":
        # If the descriptor says "last", check if this date is the last one.
        return date_obj == same_wd[-1] if same_wd else False
    else:
        # For occurrences like "1st", "2nd", etc., extract the numeric part.
        try:
            idx = int(''.join(c for c in occ if c.isdigit())) - 1
        except:
            return False
        # Check if the date is the nth occurrence.
        return date_obj == same_wd[idx] if 0 <= idx < len(same_wd) else False

# ------------------------------
# Special History: Read previous month's special-day allocations
# ------------------------------

def read_special_history(wb):
    """
    Reads the "SpecialHistory" sheet (if it exists) and returns a dictionary.
    The dictionary maps a tuple (Descriptor, EmployeeID) to an allocation count (e.g., 1 if the employee was assigned that special slot last month).
    If the sheet does not exist, returns an empty dictionary.
    """
    try:
        ws = wb["SpecialHistory"]
        data = []
        # Iterate over rows starting from row 2 (assuming row 1 is a header).
        for row in ws.iter_rows(min_row=2, values_only=True):
            data.append(row)
        # Create a DataFrame from the data.
        df = pd.DataFrame(data, columns=["Descriptor", "EmployeeID", "Allocation"])
        hist = {}
        # Populate the dictionary with keys as (Descriptor, EmployeeID).
        for _, r in df.iterrows():
            key = (r["Descriptor"], r["EmployeeID"])
            hist[key] = r["Allocation"]
        return hist
    except Exception:
        # If any error occurs (such as sheet not existing), return an empty dictionary.
        return {}

# ------------------------------
# Global ILP Rostering Solver with Fairness and Consecutive-Day Penalty
# ------------------------------

def generate_roster_schedule_ilp(excel_file, designated_min=3, big_penalty=1000,
                                 consecutive_penalty=5, fairness_coef=20):
    """
    This function builds and solves a global Integer Linear Programming (ILP) model that assigns seats to employees over the month.
    It ensures that:
      - Every employee meets their overall monthly threshold (minimum required assignments).
      - Fixed seats (pre-assigned seats) are strictly enforced.
      - For employees without fixed seats, extra flexible assignments on non-special days do not exceed their threshold.
      - Employees receive extra bonus (priority) on designated days (as defined in SubTeamOfficeDays), and a soft minimum is enforced (with slack variables) to aim for at least 'designated_min' assignments on those days.
      - A consecutive-day penalty is applied to discourage assigning the same employee on two consecutive days unless one is a designated day and the other is a special day.
      - Historical special-day allocation (from "SpecialHistory") is used to give higher priority this month to employees who were not allocated on special days in the previous month.
    The final roster is output in a new Excel sheet (named by TargetMonthYear) in a column-oriented format:
      - Column A: Employee Name
      - Row 1 (columns B onward): Dates in "YYYY-MM-DD" format.
      - Row 2 (columns B onward): The corresponding day-of-week (e.g., Mon, Tue).
      - Rows 3 onward: For each employee, the seat code assigned on that day (or blank if none).
    """
    # Load the Excel workbook and get the "Static Data" sheet.
    wb = load_workbook(excel_file)
    static_ws = wb["Static Data"]
    
    # Read the named cells: OfficePercentage and TargetMonthYear.
    office_percentage = get_named_cell_value(wb, "OfficePercentage")  # Already in [0,1], e.g., 0.6 for 60%.
    target_month_year = get_named_cell_value(wb, "TargetMonthYear")   # E.g., "Mar-25".
    month_str, year_str = target_month_year.split("-")                # Split the string into month and year parts.
    month = datetime.strptime(month_str, "%b").month                   # Convert month abbreviation to month number.
    year = int("20" + year_str) if len(year_str) == 2 else int(year_str)  # Convert year string to integer.
    
    # Read the necessary tables from the "Static Data" sheet.
    df_employees       = get_table_as_df(static_ws, "EmployeeData")       # Contains EmployeeID, EmployeeName, SubTeam, etc.
    df_seats           = get_table_as_df(static_ws, "SeatData")           # Contains SeatCode, SeatType, Days, AssignedEmployeeID, etc.
    df_public_holidays = get_table_as_df(static_ws, "PublicHolidays")     # Contains dates of public holidays.
    df_subteam_days    = get_table_as_df(static_ws, "SubTeamOfficeDays")    # Contains sub-team designated office days.
    df_special_days    = get_table_as_df(static_ws, "SpecialSubTeamDays")   # Contains special day descriptors and associated sub-teams.
    df_seat_pref       = get_table_as_df(static_ws, "SeatPreferences")      # Contains employee seat preferences.
    
    # Build lists of employee IDs and seat codes.
    employees = df_employees["EmployeeID"].tolist()
    seats = df_seats["SeatCode"].tolist()
    
    # Determine the list of working dates for the month (excluding weekends and public holidays).
    working_dates = get_working_dates(year, month, df_public_holidays["Date"])
    total_wd = len(working_dates)  # Total number of working days.
    
    # Calculate the monthly required assignments for each employee.
    req_days = {e: round(total_wd * office_percentage) for e in employees}
    
    # Create mappings for employee names and their sub-team memberships.
    emp_names = dict(zip(df_employees["EmployeeID"], df_employees["EmployeeName"]))
    emp_subteam = dict(zip(df_employees["EmployeeID"], df_employees["SubTeam"]))
    
    # Determine designated days for each employee based on their sub-team.
    # The designated_map maps a sub-team to a set of day tokens (e.g., {"mon", "wed"}).
    designated_map = {}
    for _, row in df_subteam_days.iterrows():
        st = row["SubTeam"]
        days_set = parse_days_string(row["OfficeDays"])
        designated_map.setdefault(st, set()).update(days_set)
    # For each employee, designated_days[e] is the set of working dates that are designated for their sub-team.
    designated_days = {}
    for e in employees:
        st = emp_subteam[e]
        if st in designated_map:
            designated_days[e] = {d for d in working_dates if (d.strftime("%a").lower() in designated_map[st] or d.strftime("%A").lower() in designated_map[st])}
        else:
            designated_days[e] = set()
    
    # Determine special days from the SpecialSubTeamDays table.
    # For each working date, if it matches a special descriptor, record the associated sub-team and descriptor.
    special = {}  # special[d] = (SpecialSubTeam, Descriptor)
    for d in working_dates:
        for _, row in df_special_days.iterrows():
            desc = str(row["DayDescriptor"]).strip()
            if is_day_descriptor_match(d, desc, working_dates):
                special[d] = (row["SubTeam"], desc)
                break  # Only one special sub-team per day is assumed.
    
    # Read historical special allocation from the "SpecialHistory" sheet for fairness.
    hist = read_special_history(wb)  # This returns a dictionary with keys (Descriptor, EmployeeID).
    
    # Set up seat preference bonus: if an employee prefers a seat, they get extra bonus.
    pref_bonus = {}
    for _, row in df_seat_pref.iterrows():
        e = row["EmployeeID"]
        s = row["SeatCode"]
        pref_bonus[(e, s)] = 10   # You can adjust this bonus value as needed.
    
    # Define other bonus parameters:
    fill_bonus = 1             # Bonus for simply having a seat assigned.
    designated_bonus = 5       # Bonus if assignment occurs on a designated day.
    special_bonus = 20         # Bonus if assignment occurs on a special day for the employee's sub-team.
    
    # For special days, we also add a fairness bonus if the employee did NOT get that special slot last month.
    # The fairness bonus is calculated as fairness_coef if no historical record is found for that (descriptor, employee).
    def fairness_bonus(e, desc):
        return fairness_coef if hist.get((desc, e), 0) == 0 else 0
    
    # Determine seat availability for each seat and day.
    # This checks if a seat is available on a day based on the "Days" field in SeatData.
    seat_avail = {}
    for _, srow in df_seats.iterrows():
        s_code = srow["SeatCode"]
        avail_set = parse_days_string(srow["Days"])
        for d in working_dates:
            d_abbr = d.strftime("%a").lower()
            d_full = d.strftime("%A").lower()
            seat_avail[(s_code, d)] = (d_abbr in avail_set or d_full in avail_set)
    
    # Process fixed seat assignments: for seats with type "fixed" and an AssignedEmployeeID, force the assignment.
    fixed = {}
    for _, row in df_seats.iterrows():
        if str(row["SeatType"]).strip().lower() == "fixed":
            assigned = row.get("AssignedEmployeeID")
            if pd.notna(assigned):
                s_code = row["SeatCode"]
                avail_set = parse_days_string(row["Days"])
                for d in working_dates:
                    d_abbr = d.strftime("%a").lower()
                    d_full = d.strftime("%A").lower()
                    if d_abbr in avail_set or d_full in avail_set:
                        fixed[(s_code, d)] = assigned
    
    # ------------------------------
    # Build the ILP Model using PuLP
    # ------------------------------
    
    # Create a new linear programming problem with the goal of maximizing our objective.
    model = LpProblem("GlobalTeamRostering", LpMaximize)
    
    # Decision variables: x[e,s,d] is a binary variable that is 1 if employee e is assigned seat s on day d.
    x = {}
    for e in employees:
        for s in seats:
            for d in working_dates:
                var_name = f"x_{e}_{s}_{d.strftime('%d')}"
                x[(e, s, d)] = LpVariable(var_name, cat=LpBinary)
    
    # For each employee with designated days, create a slack variable z_e (an integer >= 0)
    # This slack variable allows us to "softly" enforce a minimum number of designated-day assignments.
    z = {}
    for e in employees:
        if designated_days[e]:
            z[e] = LpVariable(f"z_{e}", lowBound=0, cat=LpInteger)
    
    # Auxiliary variables for consecutive-day assignments.
    # For each employee and each consecutive day pair, y[e,d] is 1 if the employee is assigned on both days.
    y = {}
    for e in employees:
        for i in range(len(working_dates) - 1):
            d = working_dates[i]
            # d_next is the day immediately after d in the working_dates list.
            y[(e, d)] = LpVariable(f"y_{e}_{d.strftime('%d')}", cat=LpBinary)
    
    # ---- Add Constraints to the Model ----
    
    # Constraint 1: Each seat can be occupied by at most one employee on any given day.
    for s in seats:
        for d in working_dates:
            model += lpSum(x[(e, s, d)] for e in employees) <= 1, f"SeatOccupancy_{s}_{d.strftime('%Y%m%d')}"
    
    # Constraint 2: Each employee can be assigned at most one seat per day.
    for e in employees:
        for d in working_dates:
            model += lpSum(x[(e, s, d)] for s in seats) <= 1, f"EmployeeOneSeat_{e}_{d.strftime('%Y%m%d')}"
    
    # Constraint 3: Each employee must meet or exceed their overall monthly quota (required days).
    for e in employees:
        model += lpSum(x[(e, s, d)] for s in seats for d in working_dates) >= req_days[e], f"RequiredDays_{e}"
    
    # Constraint 4: For employees with designated days, enforce at least 'designated_min' assignments on those days (using slack variable z_e).
    for e in employees:
        if designated_days[e]:
            model += lpSum(x[(e, s, d)] for s in seats for d in designated_days[e]) + z[e] >= designated_min, f"DesignatedMin_{e}"
    
    # Constraint 5: Fixed seat assignments must be enforced.
    # For each fixed seat, force the assignment for the specified employee and disallow others.
    for (s, d), e_fixed in fixed.items():
        model += x[(e_fixed, s, d)] == 1, f"FixedSeat_{s}_{d.strftime('%Y%m%d')}"
        for e in employees:
            if e != e_fixed:
                model += x[(e, s, d)] == 0, f"FixedSeatZero_{e}_{s}_{d.strftime('%Y%m%d')}"
    
    # Constraint 6: A seat cannot be assigned on a day when it is not available.
    for s in seats:
        for d in working_dates:
            if not seat_avail[(s, d)]:
                for e in employees:
                    model += x[(e, s, d)] == 0, f"SeatNotAvail_{e}_{s}_{d.strftime('%Y%m%d')}"
    
    # Constraint 7: For employees without fixed seats, do not allow extra flexible assignments on non-special days beyond their threshold.
    F = [e for e in employees if e not in set(fixed.values())]  # F is the list of employees with no fixed-seat assignments.
    for e in F:
        # non_special_days: working days that are NOT special for the employee's sub-team.
        non_special_days = [d for d in working_dates if not (d in special and emp_subteam[e] == special[d][0])]
        model += lpSum(x[(e, s, d)] for s in seats for d in non_special_days) <= req_days[e], f"NonSpecialUpper_{e}"
    
    # Constraint 8: Linearize consecutive-day assignments.
    # For each employee and consecutive working day pair, ensure that the binary variable y[e,d] reflects if the employee is assigned on both days.
    for e in employees:
        for i in range(len(working_dates) - 1):
            d = working_dates[i]
            d_next = working_dates[i+1]
            a_ed = lpSum(x[(e, s, d)] for s in seats)      # a(e,d): 1 if assigned on day d.
            a_e_next = lpSum(x[(e, s, d_next)] for s in seats)  # a(e,d_next): 1 if assigned on day d_next.
            model += y[(e, d)] <= a_ed, f"Consec1_{e}_{d.strftime('%Y%m%d')}"
            model += y[(e, d)] <= a_e_next, f"Consec2_{e}_{d.strftime('%Y%m%d')}"
            model += y[(e, d)] >= a_ed + a_e_next - 1, f"Consec3_{e}_{d.strftime('%Y%m%d')}"
    
    # Constraint 9: Avoid assigning the same employee on two consecutive days if not preferred.
    # We allow consecutive assignments if one day is designated and the other is a special day (for the employee's sub-team).
    allowed_consec = {}
    for e in employees:
        for i in range(len(working_dates) - 1):
            d = working_dates[i]
            d_next = working_dates[i+1]
            # The condition below is True if the consecutive assignment is allowed.
            cond = ((d in designated_days[e] and d_next in special and emp_subteam[e] == special[d_next][0]) or
                    (d in special and d_next in designated_days[e] and emp_subteam[e] == special[d][0]))
            allowed_consec[(e, d)] = 1 if cond else 0
    # The penalty for disallowed consecutive assignments will be added to the objective.
    
    # ---- Build the Objective Function ----
    obj_terms = []
    # Loop over every potential assignment.
    for e in employees:
        for s in seats:
            for d in working_dates:
                bonus = fill_bonus  # Start with the basic fill bonus.
                # Add bonus if there is a seat preference.
                if (e, s) in pref_bonus:
                    bonus += pref_bonus[(e, s)]
                # Add bonus if the day is designated for the employee.
                if d in designated_days[e]:
                    bonus += designated_bonus
                # If the day is special and the employee is in that special sub-team, add special and fairness bonus.
                if d in special and emp_subteam[e] == special[d][0]:
                    bonus += special_bonus + fairness_coef * (1 if (special[d][1], e) not in hist or hist.get((special[d][1], e), 0) == 0 else 0)
                # Multiply the bonus by the decision variable (which is 1 if the assignment is made, else 0).
                obj_terms.append(bonus * x[(e, s, d)])
    # Penalty for slack in designated-day assignments.
    penalty_terms = []
    for e in z:
        penalty_terms.append(big_penalty * z[e])
    # Penalty for disallowed consecutive-day assignments.
    consec_penalty_terms = []
    for e in employees:
        for i in range(len(working_dates) - 1):
            d = working_dates[i]
            if allowed_consec[(e, d)] == 0:
                consec_penalty_terms.append(consecutive_penalty * y[(e, d)])
    # The total objective is to maximize the sum of bonuses minus the penalties.
    model += lpSum(obj_terms) - lpSum(penalty_terms) - lpSum(consec_penalty_terms), "TotalObjective"
    
    # ------------------------------
    # Solve the ILP
    # ------------------------------
    model.solve()
    if LpStatus[model.status] != "Optimal":
        print("No feasible solution found or solver did not converge to an optimal solution.")
        return
    
    # ------------------------------
    # Extract the solution: for each employee and day, determine the seat assigned (if any).
    # ------------------------------
    emp_day_assign = {(e, d): None for e in employees for d in working_dates}
    for e in employees:
        for d in working_dates:
            for s in seats:
                if x[(e, s, d)].varValue == 1:
                    emp_day_assign[(e, d)] = s
                    break  # Since an employee can only be assigned one seat per day.
    
    # ------------------------------
    # Update the SpecialHistory sheet for fairness in future allocations.
    # ------------------------------
    # This sheet records, for each special day descriptor and employee, that the employee received that slot.
    if "SpecialHistory" in wb.sheetnames:
        sh_ws = wb["SpecialHistory"]
    else:
        sh_ws = wb.create_sheet("SpecialHistory")
        sh_ws.append(["Descriptor", "EmployeeID", "MonthYear"])  # Header row.
    # For every working day that is special, record which employee was assigned if they are in the special sub-team.
    for d in working_dates:
        if d in special:
            desc = special[d][1]  # The descriptor (e.g., "1st tuesday")
            for e in employees:
                if emp_day_assign[(e, d)] is not None and emp_subteam[e] == special[d][0]:
                    sh_ws.append([desc, e, target_month_year])
    
    # ------------------------------
    # Build the Output Sheet in Column-Oriented Format
    # ------------------------------
    # The output sheet will have:
    #   - Column A: "Employee Name"
    #   - Row 1 (columns B onward): Each working date (formatted as YYYY-MM-DD)
    #   - Row 2 (columns B onward): The day of the week for that date (e.g., Mon, Tue)
    #   - Rows 3 onward: Each employee's row with the seat code assigned on that day (or blank if none).
    out_sheet_name = target_month_year
    if out_sheet_name in wb.sheetnames:
        out_ws = wb[out_sheet_name]
        # Clear existing content.
        for row in out_ws.iter_rows(min_row=1, max_row=out_ws.max_row, min_col=1, max_col=out_ws.max_column):
            for cell in row:
                cell.value = None
    else:
        out_ws = wb.create_sheet(title=out_sheet_name)
    # Write header in cell A1.
    out_ws.cell(row=1, column=1).value = "Employee Name"
    # Write working dates in row 1 starting from column B.
    for j, d in enumerate(working_dates, start=2):
        out_ws.cell(row=1, column=j).value = d.strftime("%Y-%m-%d")
        out_ws.cell(row=2, column=j).value = d.strftime("%a")  # Write the day-of-week in row 2.
    out_ws.cell(row=2, column=1).value = ""
    # Write each employee's name in column A (starting from row 3) and their assigned seat code for each date.
    for i, e in enumerate(employees, start=3):
        out_ws.cell(row=i, column=1).value = emp_names[e]
        for j, d in enumerate(working_dates, start=2):
            seat_code = emp_day_assign.get((e, d), "")
            out_ws.cell(row=i, column=j).value = seat_code if seat_code else ""
    
    # Format the header rows (make them bold and fill with color).
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    for cell in out_ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for cell in out_ws[2]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    # Adjust column widths based on the maximum length of content in each column.
    for col in out_ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value and isinstance(cell.value, str):
                max_len = max(max_len, len(cell.value))
        out_ws.column_dimensions[col_letter].width = max_len + 2
    
    # Save the updated Excel workbook.
    wb.save(excel_file)
    print("Global ILP roster generated successfully.")

# ------------------------------
# Main Execution
# ------------------------------
if __name__ == "__main__":
    excel_filename = "TeamRoster.xlsx"  # Update this path if necessary.
    # Call the ILP solver function with desired parameters.
    generate_roster_schedule_ilp(excel_filename, designated_min=3, big_penalty=1000,
                                   consecutive_penalty=5, fairness_coef=20)